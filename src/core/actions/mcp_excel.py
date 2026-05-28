"""mcp_excel tool — registered via @tool decorator.

Spreadsheet reading via openpyxl and built-in csv module.

Actions:
- ``action="read" path="data/file.xlsx" sheet=0 rows=50`` — read sheet data.
- ``action="info" path="data/file.xlsx"`` — sheet names, row counts, file size.
- ``action="search" path="data/file.xlsx" query="iPhone" limit=10`` — search across cells.
- ``action="csv" path="data/file.csv" rows=50`` — read CSV file.

Path must be within ``settings.data_dir``.  openpyxl imported lazily.
"""

from __future__ import annotations

import asyncio
import csv
import logging
from pathlib import Path
from typing import Any

from src.core.actions.mcp_tools import _safe_resolve
from src.core.actions.tool_registry import tool

logger = logging.getLogger(__name__)

# ── Constants ─────────────────────────────────────────────────────────────

_MAX_ROWS = 10_000  # absolute safety limit
_DEFAULT_ROWS = 50


# ══════════════════════════════════════════════════════════════════════════
# Tool: mcp_excel
# ══════════════════════════════════════════════════════════════════════════


@tool(
    name="mcp_excel",
    description=(
        "Read spreadsheets (XLSX via openpyxl) and CSV files.\n\n"
        "Actions:\n"
        "- **read** — read a sheet by index (0-based); return first N rows.\n"
        "- **info** — sheet names, row counts per sheet, file size.\n"
        "- **search** — search across all cells for a query string.\n"
        "- **csv** — read a CSV file (same format as read).\n\n"
        "Examples:\n"
        '  action="read" path="data/file.xlsx" sheet=0 rows=50\n'
        '  action="info" path="data/file.xlsx"\n'
        '  action="search" path="data/file.xlsx" query="iPhone" limit=10\n'
        '  action="csv" path="data/file.csv" rows=50'
    ),
    category="utility",
    risk="low",
    params={
        "action": "str — 'read', 'info', 'search', or 'csv'",
        "path": "str — relative path inside data/ to the file (required)",
        "sheet": "int — sheet index (0-based, default 0, used with 'read')",
        "rows": "int — rows to return (default 50, max 10000, used with 'read' and 'csv')",
        "query": "str — search query (used with 'search')",
        "limit": "int — max search results (default 10, used with 'search')",
    },
)
async def mcp_excel(
    action: str = "",
    path: str = "",
    sheet: int = 0,
    rows: int = _DEFAULT_ROWS,
    query: str = "",
    limit: int = 10,
    **kwargs: Any,
) -> dict[str, Any]:
    """Spreadsheet and CSV reading tool."""
    try:
        if action not in ("read", "info", "search", "csv"):
            return {
                "error": (
                    f"Unknown action {action!r}. Valid actions: read, info, search, csv"
                )
            }

        if not path or not path.strip():
            return {"error": "path parameter is required"}

        resolved = _safe_resolve(path.strip())
        if resolved is None:
            return {
                "error": (
                    f"Path {path!r} is outside allowed directories or contains '..'"
                )
            }
        if not resolved.is_file():
            return {"error": f"File not found: {resolved}"}

        if action == "csv":
            return await _read_csv(resolved, max(1, min(rows, _MAX_ROWS)))
        elif action == "info":
            return await _excel_info(resolved)
        elif action == "search":
            if not query or not query.strip():
                return {"error": "query parameter is required for action='search'"}
            return await _excel_search(resolved, query.strip(), max(1, limit))
        else:  # read
            return await _excel_read(
                resolved, max(0, sheet), max(1, min(rows, _MAX_ROWS))
            )
    except ValueError as exc:
        return {"error": str(exc)}
    except Exception as exc:
        logger.exception("mcp_excel(%r) failed", action)
        return {"error": f"Unexpected error: {exc}"}


# ══════════════════════════════════════════════════════════════════════════
# Helpers — openpyxl lazy import
# ══════════════════════════════════════════════════════════════════════════


def _import_openpyxl() -> Any:
    """Lazy import of openpyxl. Raises ImportError if not installed."""
    try:
        import openpyxl  # type: ignore[import-untyped]

        return openpyxl
    except ImportError:
        raise ImportError("openpyxl not installed: pip install openpyxl")


# ══════════════════════════════════════════════════════════════════════════
# Action implementations — XLSX
# ══════════════════════════════════════════════════════════════════════════


async def _excel_read(file_path: Path, sheet_idx: int, max_rows: int) -> dict[str, Any]:
    """Read a sheet by index (0-based)."""
    openpyxl = _import_openpyxl()
    loop = asyncio.get_running_loop()

    def _do() -> dict[str, Any]:
        try:
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"Cannot open workbook: {exc}")

        sheet_names = wb.sheetnames
        if sheet_idx >= len(sheet_names):
            wb.close()
            return {
                "error": (
                    f"Sheet index {sheet_idx} out of range. "
                    f"Available sheets (0-{len(sheet_names) - 1}): {sheet_names}"
                )
            }

        ws = wb[sheet_names[sheet_idx]]
        data: list[list[Any]] = []
        row_count = 0
        col_count = 0

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            # Convert row — merged cells yield None, handle gracefully
            processed: list[Any] = []
            for cell in row:
                if cell is None:
                    processed.append(None)
                elif isinstance(cell, (int, float)):
                    processed.append(cell)
                else:
                    processed.append(str(cell))
            data.append(processed)
            col_count = max(col_count, len(processed))
            row_count += 1

        wb.close()

        return {
            "ok": True,
            "action": "read",
            "sheet_name": sheet_names[sheet_idx],
            "sheet_index": sheet_idx,
            "rows_returned": row_count,
            "columns_detected": col_count,
            "data": data,
        }

    try:
        result = await loop.run_in_executor(None, _do)
    except ImportError as exc:
        return {"error": str(exc)}
    except ValueError as exc:
        return {"error": str(exc)}
    except Exception as exc:
        logger.warning("Excel read error: %s", exc)
        return {"error": f"Failed to read spreadsheet: {exc}"}

    return result


async def _excel_info(file_path: Path) -> dict[str, Any]:
    """Get spreadsheet metadata: sheet names, row counts, file size."""
    openpyxl = _import_openpyxl()
    loop = asyncio.get_running_loop()

    def _do() -> dict[str, Any]:
        try:
            wb = openpyxl.load_workbook(str(file_path), read_only=True)
        except Exception as exc:
            raise ValueError(f"Cannot open workbook: {exc}")

        sheet_names = wb.sheetnames
        sheets_info: list[dict[str, Any]] = []
        for name in sheet_names:
            ws = wb[name]
            try:
                row_count = ws.max_row or 0
                col_count = ws.max_column or 0
            except Exception:
                row_count = 0
                col_count = 0
            sheets_info.append(
                {
                    "name": name,
                    "rows": row_count,
                    "columns": col_count,
                }
            )

        wb.close()

        file_size = file_path.stat().st_size
        return {
            "ok": True,
            "action": "info",
            "sheets": sheets_info,
            "sheet_count": len(sheets_info),
            "file_size_bytes": file_size,
            "file_size_kb": round(file_size / 1024, 1),
        }

    try:
        result = await loop.run_in_executor(None, _do)
    except ImportError as exc:
        return {"error": str(exc)}
    except ValueError as exc:
        return {"error": str(exc)}
    except Exception as exc:
        logger.warning("Excel info error: %s", exc)
        return {"error": f"Failed to read spreadsheet info: {exc}"}

    return result


async def _excel_search(file_path: Path, query: str, limit: int) -> dict[str, Any]:
    """Search across all sheets for a query string (case-insensitive)."""
    openpyxl = _import_openpyxl()
    loop = asyncio.get_running_loop()

    query_lower = query.lower()

    def _do() -> dict[str, Any]:
        try:
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"Cannot open workbook: {exc}")

        results: list[dict[str, Any]] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    if len(results) >= limit:
                        break
                    if cell is not None:
                        cell_str = str(cell)
                        if query_lower in cell_str.lower():
                            results.append(
                                {
                                    "sheet": sheet_name,
                                    "row": row_idx,
                                    "column": col_idx,
                                    "value": cell_str,
                                }
                            )
                    if len(results) >= limit:
                        break
                if len(results) >= limit:
                    break
            if len(results) >= limit:
                break

        wb.close()

        return {
            "ok": True,
            "action": "search",
            "query": query,
            "results": results,
            "total_found": len(results),
            "limit": limit,
            "truncated": len(results) >= limit,
        }

    try:
        result = await loop.run_in_executor(None, _do)
    except ImportError as exc:
        return {"error": str(exc)}
    except ValueError as exc:
        return {"error": str(exc)}
    except Exception as exc:
        logger.warning("Excel search error: %s", exc)
        return {"error": f"Failed to search spreadsheet: {exc}"}

    return result


# ══════════════════════════════════════════════════════════════════════════
# Action implementations — CSV
# ══════════════════════════════════════════════════════════════════════════


async def _read_csv(file_path: Path, max_rows: int) -> dict[str, Any]:
    """Read a CSV file using the built-in csv module.

    Tries UTF-8 first, then falls back to cp1251 for legacy Windows files.
    """
    loop = asyncio.get_running_loop()

    def _do() -> dict[str, Any]:
        raw_bytes = file_path.read_bytes()
        encoding = "utf-8"

        # Try UTF-8 (with BOM support) first, then cp1251
        try:
            content = raw_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                content = raw_bytes.decode("cp1251")
                encoding = "cp1251"
            except UnicodeDecodeError:
                content = raw_bytes.decode("utf-8", errors="replace")
                encoding = "utf-8 (with replacements)"

        reader = csv.reader(content.splitlines())
        data: list[list[str]] = []
        row_count = 0
        col_count = 0

        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            # Skip completely empty rows
            if any(cell.strip() for cell in row):
                data.append(row)
                col_count = max(col_count, len(row))
                row_count += 1

        return {
            "ok": True,
            "action": "csv",
            "encoding": encoding,
            "rows_returned": row_count,
            "columns_detected": col_count,
            "data": data,
        }

    try:
        result = await loop.run_in_executor(None, _do)
    except ValueError as exc:
        return {"error": str(exc)}
    except Exception as exc:
        logger.warning("CSV read error: %s", exc)
        return {"error": f"Failed to read CSV: {exc}"}

    return result
